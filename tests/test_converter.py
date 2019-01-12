import csv
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from converter import Converter


@pytest.fixture(scope='module')
def file_path():
    d = {
        'Test': 'Hi',
        'Test list': ['Hello', 'Bye bye'],
        'Numbers!': [1, 2, 3],
        'Hell': 'Yeah!',
        'Will this work': {
            'Hey': [1, 2]
        }
    }
    p = Path('test.json')
    with open(p, 'w') as fp:
        json.dump(d, fp)

    yield p

    files = [Path('result.csv'), Path('result.xlsx'), p]
    for file in files:
        try:
            file.unlink()
        except FileNotFoundError:
            continue


@pytest.fixture(scope='module')
def file_path_help():
    d = {
        'Test': 'Hi',
        'Numbers!': [1, 2, 3],
        'Hell': 'Yeah!',
        'Will this work': {
            'Hey': [1, 2]
        },
        'Dict help': {'key': 'key', 'value': 1, 'random_value': 'woop'}
    }
    p = Path('test.json')
    with open(p, 'w') as fp:
        json.dump(d, fp)

    yield p

    files = [Path('result.csv'), Path('result.xlsx'), p]
    for file in files:
        try:
            file.unlink()
        except FileNotFoundError:
            continue


def test_export_to_csv(file_path):
    with Converter(file_path, 'csv', 'result') as c:
        c.convert()

    with open('result.csv', 'r') as fp:
        csv_reader = csv.reader(fp, delimiter=',')

        line_count = 0
        for row in csv_reader:
            if line_count == 0:
                assert row == ['Test', 'Test list', 'Numbers!', 'Hell', 'Will this work_Hey']
            if line_count == 1:
                assert row == ['Hi', 'Hello', '1', 'Yeah!', '1']  # saving loses type
            if line_count == 2:
                assert row == ['', 'Bye bye', '2', '', '2']
            if line_count == 3:
                assert row == ['', '', '3', '', '']

            line_count += 1


def test_export_to_xlsx(file_path):
    with Converter(file_path, 'xlsx', 'result') as c:
        c.convert()

    wb = load_workbook(filename='result.xlsx')
    ws = wb.active

    line_count = 0
    for c1, c2, c3, c4, c5 in ws[ws.dimensions]:
        row = [c1.value, c2.value, c3.value, c4.value, c5.value]

        if line_count == 0:
            assert row == ['Test', 'Test list', 'Numbers!', 'Hell', 'Will this work_Hey']
        if line_count == 1:
            assert row == ['Hi', 'Hello', 1, 'Yeah!', 1]  # xlsx keeps numbers
        if line_count == 2:
            assert row == [None, 'Bye bye', 2, None, 2]
        if line_count == 3:
            assert row == [None, None, 3, None, None]

        line_count += 1


def test_export_to_csv_with_helper(file_path_help):
    helper = {
        'key': 'key',
        'value': 'value'
    }
    with Converter(file_path_help, 'csv', 'result') as c:
        c.convert(helper=helper)

    with open('result.csv', 'r') as fp:
        csv_reader = csv.reader(fp, delimiter=',')

        line_count = 0
        for row in csv_reader:
            if line_count == 0:
                assert row == [
                    'Test', 'Numbers!', 'Hell', 'Will this work_Hey',
                    'Dict help_key'
                ]
            if line_count == 1:
                assert row == ['Hi', '1', 'Yeah!', '1', '1']
            if line_count == 2:
                assert row == ['', '2', '', '2', '']
            if line_count == 3:
                assert row == ['', '3', '', '', '']

            line_count += 1


def test_export_to_xlsx_with_helper(file_path_help):
    helper = {
        'key': 'key',
        'value': 'value'
    }
    with Converter(file_path_help, 'xlsx', 'result') as c:
        c.convert(helper=helper)

    wb = load_workbook(filename='result.xlsx')
    ws = wb.active

    line_count = 0
    for c1, c2, c3, c4, c5 in ws[ws.dimensions]:
        row = [c1.value, c2.value, c3.value, c4.value, c5.value]

        if line_count == 0:
            assert row == [
                'Test', 'Numbers!', 'Hell', 'Will this work_Hey',
                'Dict help_key'
            ]
        if line_count == 1:
            assert row == ['Hi', 1, 'Yeah!', 1, 1]
        if line_count == 2:
            assert row == [None, 2, None, 2, None]
        if line_count == 3:
            assert row == [None, 3, None, None, None]

        line_count += 1
